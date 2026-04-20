"""거래내역 엑셀 임포트 — 템플릿 다운로드 / 미리보기 / 저장"""
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from pydantic import BaseModel, Field
from sqlalchemy.orm import Session

from .. import models
from ..database import get_db

router = APIRouter(prefix="/api/imports", tags=["imports"])

TEMPLATE_COLUMNS = ["날짜", "구분", "내용", "금액", "카테고리", "메모"]
TYPE_MAP = {
    "수입": models.TransactionType.INCOME,
    "INCOME": models.TransactionType.INCOME,
    "지출": models.TransactionType.EXPENSE,
    "EXPENSE": models.TransactionType.EXPENSE,
}


# --- Schemas ---
class PreviewRow(BaseModel):
    row_number: int
    date: Optional[str] = None
    type: Optional[str] = None
    description: Optional[str] = None
    amount: Optional[str] = None
    category_id: Optional[int] = None
    category_name: Optional[str] = None
    memo: Optional[str] = None
    valid: bool = True
    errors: list[str] = Field(default_factory=list)


class PreviewResponse(BaseModel):
    total: int
    valid_count: int
    error_count: int
    rows: list[PreviewRow]


class ConfirmRow(BaseModel):
    date: str
    type: str
    description: str
    amount: str
    category_id: Optional[int] = None
    memo: Optional[str] = None


class ConfirmRequest(BaseModel):
    rows: list[ConfirmRow]


class ConfirmResponse(BaseModel):
    inserted: int


# --- Helpers ---
def _parse_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_amount(v) -> Optional[Decimal]:
    if v is None or v == "":
        return None
    try:
        s = str(v).replace(",", "").replace("원", "").strip()
        d = Decimal(s)
        return d if d > 0 else None
    except (InvalidOperation, ValueError):
        return None


def _parse_type(v) -> Optional[models.TransactionType]:
    if v is None:
        return None
    key = str(v).strip().upper() if isinstance(v, str) else str(v)
    return TYPE_MAP.get(str(v).strip()) or TYPE_MAP.get(key)


def _resolve_category(
    db: Session, name: Optional[str], ttype: Optional[models.TransactionType]
) -> tuple[Optional[int], Optional[str], Optional[str]]:
    """카테고리 이름 → id 변환. (id, 표시명, 에러) 반환"""
    if not name:
        return None, None, None
    name = str(name).strip()
    if not name:
        return None, None, None
    q = db.query(models.Category).filter(models.Category.name == name)
    if ttype:
        q = q.filter(models.Category.type == ttype)
    cat = q.first()
    if cat:
        return cat.id, cat.name, None
    return None, name, f"알 수 없는 카테고리: {name}"


# --- Endpoints ---
@router.get("/transactions/template")
def download_template():
    """엑셀 템플릿 다운로드"""
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, h in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    samples = [
        ["2026-04-01", "수입", "4월 제품매출", 3000000, "제품매출", "대기업 납품"],
        ["2026-04-02", "지출", "사무실 임차료", 800000, "임차료", ""],
        ["2026-04-03", "지출", "4월 급여", 5000000, "급여", ""],
    ]
    for r, row in enumerate(samples, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)

    widths = [12, 8, 28, 14, 14, 24]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="transactions_template.xlsx"'},
    )


@router.post("/transactions/preview", response_model=PreviewResponse)
async def preview_transactions(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """엑셀 파일을 파싱해서 미리보기 + 유효성 결과 반환 (저장하지 않음)"""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "엑셀 파일(.xlsx)만 업로드 가능합니다")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"엑셀 파일을 열 수 없습니다: {e}")

    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if not header:
        raise HTTPException(400, "빈 파일입니다")

    header_norm = [str(h).strip() if h is not None else "" for h in header]
    expected = set(TEMPLATE_COLUMNS)
    if not expected.issubset(set(header_norm)):
        raise HTTPException(
            400,
            f"헤더가 올바르지 않습니다. 필요 컬럼: {', '.join(TEMPLATE_COLUMNS)}",
        )
    idx = {col: header_norm.index(col) for col in TEMPLATE_COLUMNS}

    preview: list[PreviewRow] = []
    row_number = 1
    for raw in rows_iter:
        row_number += 1
        if raw is None or all(v is None or v == "" for v in raw):
            continue

        def cell(col):
            i = idx[col]
            return raw[i] if i < len(raw) else None

        errors: list[str] = []

        d = _parse_date(cell("날짜"))
        if d is None:
            errors.append("날짜 형식 오류 (예: 2026-04-01)")

        t = _parse_type(cell("구분"))
        if t is None:
            errors.append("구분은 '수입' 또는 '지출'이어야 합니다")

        desc_val = cell("내용")
        desc = str(desc_val).strip() if desc_val is not None else ""
        if not desc:
            errors.append("내용은 필수입니다")

        amt = _parse_amount(cell("금액"))
        if amt is None:
            errors.append("금액은 0보다 큰 숫자여야 합니다")

        cat_val = cell("카테고리")
        cat_name_raw = str(cat_val).strip() if cat_val is not None else ""
        cat_id, cat_name, cat_err = _resolve_category(db, cat_name_raw or None, t)
        if cat_err:
            errors.append(cat_err)

        memo_val = cell("메모")
        memo = str(memo_val).strip() if memo_val is not None else ""

        preview.append(
            PreviewRow(
                row_number=row_number,
                date=d.isoformat() if d else None,
                type=t.value if t else None,
                description=desc or None,
                amount=str(amt) if amt is not None else None,
                category_id=cat_id,
                category_name=cat_name,
                memo=memo or None,
                valid=len(errors) == 0,
                errors=errors,
            )
        )

    valid_count = sum(1 for r in preview if r.valid)
    return PreviewResponse(
        total=len(preview),
        valid_count=valid_count,
        error_count=len(preview) - valid_count,
        rows=preview,
    )


@router.post(
    "/transactions/confirm",
    response_model=ConfirmResponse,
    status_code=status.HTTP_201_CREATED,
)
def confirm_transactions(payload: ConfirmRequest, db: Session = Depends(get_db)):
    """미리보기에서 유효한 행만 클라이언트가 보내와 일괄 저장"""
    if not payload.rows:
        raise HTTPException(400, "저장할 행이 없습니다")

    inserted = 0
    for r in payload.rows:
        d = _parse_date(r.date)
        t = _parse_type(r.type)
        amt = _parse_amount(r.amount)
        if not (d and t and amt and r.description):
            raise HTTPException(400, f"유효하지 않은 행: {r}")
        if r.category_id and not db.get(models.Category, r.category_id):
            raise HTTPException(400, f"알 수 없는 category_id: {r.category_id}")
        db.add(
            models.Transaction(
                date=d,
                description=r.description,
                amount=amt,
                type=t,
                category_id=r.category_id,
                memo=r.memo,
            )
        )
        inserted += 1
    db.commit()
    return ConfirmResponse(inserted=inserted)
