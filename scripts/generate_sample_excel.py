"""교육용 샘플 엑셀 생성 스크립트.

12개월치(12 months) 가상 거래를 생성해 samples/transactions_sample_12months.xlsx로 저장.
실행: python scripts/generate_sample_excel.py

데이터 특성:
- 기준 종료일: 오늘. 종료일로부터 12개월 전까지 월별 데이터 생성.
- 분기별 성수기·비수기 반영 (Q4 매출 성수기, Q1 비용 집중).
- 실제 카테고리명과 매칭되도록 기본 시드와 동일한 카테고리 사용.
- 일부 "오류 케이스" 행은 포함하지 않음 — 교육용 정상 데이터셋.
"""
from __future__ import annotations

import random
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADERS = ["날짜", "구분", "내용", "금액", "카테고리", "메모"]

# 기본 시드와 동일한 카테고리 (seed.py 기준)
INCOME_CATEGORIES = ["제품매출", "서비스매출", "이자수익"]
EXPENSE_CATEGORIES = ["급여", "임차료", "광고비", "통신비", "소모품비", "식비", "교통비"]

# 월별 성수기 가중치 (1.0이 평균, 1.3은 30% 증가)
MONTHLY_REVENUE_WEIGHT = {
    1: 0.8, 2: 0.7, 3: 0.9, 4: 1.0, 5: 1.1, 6: 1.0,
    7: 0.9, 8: 0.8, 9: 1.0, 10: 1.2, 11: 1.4, 12: 1.5,  # Q4 성수기
}
MONTHLY_EXPENSE_WEIGHT = {
    1: 1.3, 2: 1.2, 3: 1.1, 4: 1.0, 5: 1.0, 6: 1.0,  # Q1 비용 집중(상여·세금)
    7: 1.0, 8: 1.0, 9: 1.0, 10: 1.0, 11: 1.1, 12: 1.2,
}


def month_range(end: date, months: int) -> list[tuple[int, int]]:
    """종료일 기준 최근 N개월의 (year, month) 리스트 반환 (오래된 → 최근)."""
    result = []
    y, m = end.year, end.month
    for _ in range(months):
        result.append((y, m))
        m -= 1
        if m == 0:
            m = 12
            y -= 1
    return list(reversed(result))


def random_day_in_month(year: int, month: int, max_day: int | None = None) -> int:
    """해당 월에서 1 ~ max_day 범위 내 무작위 일자."""
    import calendar
    last = calendar.monthrange(year, month)[1]
    upper = min(last, max_day) if max_day else last
    return random.randint(1, upper)


def generate_rows(end: date) -> list[list]:
    rows: list[list] = []
    today = end
    for (y, m) in month_range(end, 12):
        rev_w = MONTHLY_REVENUE_WEIGHT[m]
        exp_w = MONTHLY_EXPENSE_WEIGHT[m]

        # 이번 달이면 오늘 이후 날짜 생성 금지
        max_day = today.day if (y == today.year and m == today.month) else None

        # 수입: 월 3~5건
        for _ in range(random.randint(3, 5)):
            cat = random.choice(INCOME_CATEGORIES)
            base_amount = {
                "제품매출": random.randint(2000, 8000),
                "서비스매출": random.randint(1000, 4000),
                "이자수익": random.randint(50, 300),
            }[cat]
            amount = int(base_amount * rev_w) * 1000
            d = random_day_in_month(y, m, max_day)
            rows.append([
                date(y, m, d).isoformat(),
                "수입",
                f"{m}월 {cat}",
                amount,
                cat,
                "",
            ])

        # 지출: 월 8~12건
        # 고정비는 매월 반드시 포함 (임차료, 통신비, 급여)
        fixed_expenses = [
            ("임차료", random.randint(700, 900) * 1000, f"{m}월 사무실 임차료"),
            ("통신비", random.randint(50, 120) * 1000, f"{m}월 통신요금"),
            ("급여", random.randint(4000, 6000) * 1000 * 1, f"{m}월 직원 급여"),
        ]
        for cat, amount, desc in fixed_expenses:
            d = random_day_in_month(y, m, max_day)
            rows.append([
                date(y, m, d).isoformat(),
                "지출",
                desc,
                int(amount * exp_w),
                cat,
                "고정비",
            ])

        # 변동비 5~9건
        variable_cats = ["광고비", "소모품비", "식비", "교통비"]
        for _ in range(random.randint(5, 9)):
            cat = random.choice(variable_cats)
            base = {
                "광고비": random.randint(100, 800),
                "소모품비": random.randint(20, 150),
                "식비": random.randint(30, 200),
                "교통비": random.randint(10, 100),
            }[cat]
            amount = int(base * exp_w) * 1000
            d = random_day_in_month(y, m, max_day)
            rows.append([
                date(y, m, d).isoformat(),
                "지출",
                f"{m}월 {cat}",
                amount,
                cat,
                "",
            ])

    rows.sort(key=lambda r: r[0])
    return rows


def write_xlsx(rows: list[list], path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "거래내역"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2563EB")
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)

    widths = [12, 8, 32, 14, 14, 16]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main():
    random.seed(20260420)  # 재현 가능한 데이터
    end = date.today()
    rows = generate_rows(end)

    out = Path(__file__).resolve().parents[1] / "samples" / "transactions_sample_12months.xlsx"
    write_xlsx(rows, out)

    total_income = sum(r[3] for r in rows if r[1] == "수입")
    total_expense = sum(r[3] for r in rows if r[1] == "지출")
    print(f"[ok] generated {len(rows)} rows → {out}")
    print(f"     수입 합계: {total_income:,}원 / 지출 합계: {total_expense:,}원")
    print(f"     기간: {rows[0][0]} ~ {rows[-1][0]}")


if __name__ == "__main__":
    main()
